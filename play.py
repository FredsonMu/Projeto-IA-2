import time
from pathlib import Path

import pygame
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from Connect4Game import Connect4Game
from HumanPlayer import HumanPlayer
from MCTSAIPlayer import MCTSAIPlayer
from MinimaxAIPlayer import MinimaxAIPlayer

XLSX_PATH = Path("resultados.xlsx")

COLUMNS = [
    "Comparacao", "Parametros usados", "No de Jogos",
    "Vitorias Jogador 1", "Vitorias Jogador 2", "Empates",
    "Taxa de Vitorias Jogador 1", "Taxa de Vitorias Jogador 2",
    "Duracao Media do Jogo", "Duracao Maxima", "Duracao Minima",
    "Diferencas de Comportamento Observadas",
]


# ── Verbose wrappers ──────────────────────────────────────────────────────────

class _VerboseMinimax(MinimaxAIPlayer):
    def get_move(self, board):
        score = self.evaluate_board(board, self.piece)
        move = super().get_move(board)
        print(f"  [Minimax] Avaliacao: {score:+d}  |  Jogada: coluna {move}")
        return move


class _VerboseMCTS(MCTSAIPlayer):
    def get_move(self, board):
        move = super().get_move(board)
        print(f"  [MCTS]    Jogada: coluna {move}")
        return move


# ── Stats tracking ────────────────────────────────────────────────────────────

def _empty_stats(params, obs):
    return {"jogos": 0, "v_humano": 0, "v_ia": 0, "empates": 0,
            "durations": [], "params": params, "obs": obs}


# ── Excel update ──────────────────────────────────────────────────────────────

def _save_to_excel(resultados):
    if not XLSX_PATH.exists():
        print(f"  AVISO: {XLSX_PATH} nao encontrado, nao foi possivel guardar.")
        return

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    # Remove linhas placeholder de "Humano vs IA"
    rows_to_delete = [
        row[0].row for row in ws.iter_rows(min_row=2)
        if "Humano" in str(row[0].value or "")
    ]
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for label, stats in resultados.items():
        if stats["jogos"] == 0:
            continue
        jogos = stats["jogos"]
        v1 = stats["v_humano"]
        v2 = stats["v_ia"]
        e  = stats["empates"]
        dur = stats["durations"]

        ws.append([
            f"Humano vs {label}",
            stats["params"],
            jogos, v1, v2, e,
            v1 / jogos,
            v2 / jogos,
            sum(dur) / len(dur),
            max(dur),
            min(dur),
            stats["obs"],
        ])

        nr = ws.max_row
        for col in (7, 8):
            ws.cell(row=nr, column=col).number_format = "0.0%"
        for col in (9, 10, 11):
            ws.cell(row=nr, column=col).number_format = "0.000"
        for cell in ws[nr]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(XLSX_PATH)
    print(f"\n  Resultados guardados em '{XLSX_PATH}'")


# ── UI helpers ────────────────────────────────────────────────────────────────

def _mostrar_placar(scores):
    h, ia, e = scores["Humano"], scores["IA"], scores["Empate"]
    print(f"\n  Placar  ->  Humano: {h}  |  IA: {ia}  |  Empates: {e}")
    print("  " + "-" * 40)


def _escolher_ia():
    while True:
        print("\n  Escolhe a IA adversaria:")
        print("  (1) Minimax  (depth=4,  ~86 ms/jogada)")
        print("  (2) MCTS     (500 iter, ~440 ms/jogada)")
        opcao = input("  Opcao: ").strip()
        if opcao == "1":
            return (_VerboseMinimax(piece=2, max_depth=4), "Minimax")
        if opcao == "2":
            return (_VerboseMCTS(piece=2, max_iterations=500), "MCTS")
        print("  Opcao invalida, tenta de novo.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n  *** Connect 4 — Humano vs IA ***")
    print("  Humano = Vermelho  (Jogador 1)")
    print("  IA     = Amarelo   (Jogador 2)")
    print("  Clica numa coluna para jogar.\n")

    placar   = {"Humano": 0, "IA": 0, "Empate": 0}
    resultados = {
        "Minimax": _empty_stats(
            params="Humano vs Minimax max_depth=4",
            obs="Jogos manuais. Humano (Vermelho, J1) vs Minimax com poda alfa-beta (Amarelo, J2).",
        ),
        "MCTS": _empty_stats(
            params="Humano vs MCTS max_iterations=500",
            obs="Jogos manuais. Humano (Vermelho, J1) vs MCTS com UCB1 (Amarelo, J2).",
        ),
    }

    while True:
        _mostrar_placar(placar)
        ia_player, ia_label = _escolher_ia()
        print(f"\n  A jogar contra: {ia_label}")

        p1   = HumanPlayer(piece=1)
        game = Connect4Game()

        t0     = time.perf_counter()
        winner = game.run_game(p1, ia_player, headless=False)
        dur    = time.perf_counter() - t0

        # Fecha janela 5s apos terminar (game_over ja espera 3s + 2s aqui)
        pygame.time.wait(2000)
        pygame.quit()

        # Registar resultado
        stats = resultados[ia_label]
        stats["jogos"]    += 1
        stats["durations"].append(dur)

        if winner == 1:
            placar["Humano"]  += 1
            stats["v_humano"] += 1
            print("\n  Resultado: GANHOU O HUMANO!")
        elif winner == 2:
            placar["IA"]  += 1
            stats["v_ia"] += 1
            print("\n  Resultado: GANHOU A IA!")
        else:
            placar["Empate"]  += 1
            stats["empates"]  += 1
            print("\n  Resultado: EMPATE!")

        resposta = input("\n  Jogar novamente? (s/n): ").strip().lower()
        if resposta != "s":
            break

    _mostrar_placar(placar)

    # Publicar resultados no Excel
    total = sum(s["jogos"] for s in resultados.values())
    if total > 0:
        _save_to_excel(resultados)

    print("  Obrigado por jogar!\n")


if __name__ == "__main__":
    main()
