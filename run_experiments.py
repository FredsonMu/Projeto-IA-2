import argparse
import random
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from Connect4Board import Connect4Board
from MCTSAIPlayer import MCTSAIPlayer
from MinimaxAIPlayer import MinimaxAIPlayer
from RandomPlayer import RandomAIPlayer


COLUMNS = [
    "Comparacao",
    "Parametros usados",
    "No de Jogos",
    "Vitorias Jogador 1",
    "Vitorias Jogador 2",
    "Empates",
    "Taxa de Vitorias Jogador 1",
    "Taxa de Vitorias Jogador 2",
    "Duracao Media do Jogo",
    "Duracao Maxima",
    "Duracao Minima",
    "Diferencas de Comportamento Observadas",
]


def play_game(player1, player2, rows=6, cols=7, n_connect=4):
    board = Connect4Board(rows, cols, n_connect)
    players = [player1, player2]
    turn = 0
    moves = 0
    started_at = time.perf_counter()

    while True:
        current_player = players[turn]
        valid_moves = board.get_valid_moves()

        if not valid_moves:
            winner = 0
            break

        move = current_player.get_move(board)
        if move not in valid_moves:
            raise ValueError(
                f"Jogada invalida do jogador {current_player.piece}: {move}. "
                f"Jogadas validas: {valid_moves}"
            )

        board.drop_piece(move, current_player.piece)
        moves += 1

        if board.check_winner(current_player.piece):
            winner = current_player.piece
            break

        if board.is_board_full():
            winner = 0
            break

        turn = (turn + 1) % 2

    duration = time.perf_counter() - started_at
    return winner, duration, moves


def run_match(comparison, player1_factory, player2_factory, games, parameters, observations):
    wins_player1 = 0
    wins_player2 = 0
    draws = 0
    durations = []

    for game_index in range(games):
        random.seed(1000 + game_index)
        player1 = player1_factory(piece=1, game_index=game_index)
        player2 = player2_factory(piece=2, game_index=game_index)
        winner, duration, _ = play_game(player1, player2)

        durations.append(duration)
        if winner == 1:
            wins_player1 += 1
        elif winner == 2:
            wins_player2 += 1
        else:
            draws += 1

    return {
        "Comparacao": comparison,
        "Parametros usados": parameters,
        "No de Jogos": games,
        "Vitorias Jogador 1": wins_player1,
        "Vitorias Jogador 2": wins_player2,
        "Empates": draws,
        "Taxa de Vitorias Jogador 1": wins_player1 / games,
        "Taxa de Vitorias Jogador 2": wins_player2 / games,
        "Duracao Media do Jogo": sum(durations) / len(durations),
        "Duracao Maxima": max(durations),
        "Duracao Minima": min(durations),
        "Diferencas de Comportamento Observadas": observations,
    }


def write_results_xlsx(rows, output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"
    sheet.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        sheet.append([row.get(column, "") for column in COLUMNS])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, column_name in enumerate(COLUMNS, start=1):
        column_letter = get_column_letter(column_index)
        width = 18
        if column_name in ("Comparacao", "Parametros usados"):
            width = 28
        elif column_name == "Diferencas de Comportamento Observadas":
            width = 55
        sheet.column_dimensions[column_letter].width = width

    for row in range(2, sheet.max_row + 1):
        for column in (7, 8):
            sheet.cell(row=row, column=column).number_format = "0.0%"
        for column in (9, 10, 11):
            sheet.cell(row=row, column=column).number_format = "0.000"

    workbook.save(output_path)


def placeholder(comparison, note):
    return {
        "Comparacao": comparison,
        "Parametros usados": "A preencher",
        "Diferencas de Comportamento Observadas": note,
    }


def main():
    parser = argparse.ArgumentParser(description="Corre testes automaticos e gera resultados.xlsx.")
    parser.add_argument("--games", type=int, default=20)
    parser.add_argument("--mcts-iterations", type=int, default=250)
    parser.add_argument("--output", default="resultados.xlsx")
    args = parser.parse_args()

    rows = []

    # ── Minimax vs Aleatorio ──────────────────────────────────────────────
    print("A correr: Minimax vs Aleatorio...")
    minimax_random_row = run_match(
        comparison="Minimax vs Aleatorio",
        player1_factory=lambda piece, game_index: MinimaxAIPlayer(piece=piece, max_depth=4),
        player2_factory=lambda piece, game_index: RandomAIPlayer(piece=piece),
        games=args.games,
        parameters="Minimax max_depth=4; Random sem parametros",
        observations=(
            "O Minimax avalia sistematicamente a arvore de jogo ate profundidade 4 e "
            "escolhe sempre a jogada com maior valor heuristico. O aleatorio nao planeia "
            "e perde quase sempre contra qualquer agente que calcule ameacas."
        ),
    )
    rows.append(minimax_random_row)
    print(f"  -> J1: {minimax_random_row['Vitorias Jogador 1']}  J2: {minimax_random_row['Vitorias Jogador 2']}  Empates: {minimax_random_row['Empates']}  Media: {minimax_random_row['Duracao Media do Jogo']:.3f}s")

    # ── MCTS vs Aleatorio ─────────────────────────────────────────────────
    print("A correr: MCTS vs Aleatorio...")
    mcts_random_row = run_match(
        comparison="MCTS vs Aleatorio",
        player1_factory=lambda piece, game_index: MCTSAIPlayer(
            piece=piece,
            max_iterations=args.mcts_iterations,
            random_seed=2000 + game_index,
        ),
        player2_factory=lambda piece, game_index: RandomAIPlayer(piece=piece),
        games=args.games,
        parameters=f"MCTS max_iterations={args.mcts_iterations}; Random sem parametros",
        observations=(
            "MCTS escolhe jogadas com base em simulacoes aleatorias e tende a melhorar "
            "quando se aumenta max_iterations. O aleatorio nao planeia e escolhe apenas "
            "uma coluna valida ao acaso."
        ),
    )
    rows.append(mcts_random_row)
    print(f"  -> J1: {mcts_random_row['Vitorias Jogador 1']}  J2: {mcts_random_row['Vitorias Jogador 2']}  Empates: {mcts_random_row['Empates']}  Media: {mcts_random_row['Duracao Media do Jogo']:.3f}s")

    # ── Minimax vs MCTS — 3 combinacoes (tempos calibrados por jogada) ───
    # Medicoes: depth=3≈21ms, depth=4≈86ms, depth=5≈380ms por jogada
    #           iter=25≈19ms, iter=100≈86ms, iter=430≈380ms por jogada
    combos = [
        (3,  25,  "1a comb Minimax vs MCTS"),
        (4, 100,  "2a comb Minimax vs MCTS"),
        (5, 430,  "3a comb Minimax vs MCTS"),
    ]

    for depth, iterations, label in combos:
        print(f"A correr: {label} (depth={depth}, iter={iterations})...")
        # capture loop variables
        d, it = depth, iterations
        row = run_match(
            comparison=label,
            player1_factory=lambda piece, game_index, d=d: MinimaxAIPlayer(piece=piece, max_depth=d),
            player2_factory=lambda piece, game_index, it=it: MCTSAIPlayer(
                piece=piece,
                max_iterations=it,
                random_seed=3000 + game_index,
            ),
            games=args.games,
            parameters=f"Minimax max_depth={d}; MCTS max_iterations={it}",
            observations=(
                f"Parametros calibrados para tempo por jogada equivalente (~{[21,86,380][[3,4,5].index(d)]}ms). "
                f"Minimax (profundidade {d}) usa poda alfa-beta e heuristica posicional. "
                f"MCTS ({it} iteracoes) usa simulacoes aleatorias com UCB1."
            ),
        )
        rows.append(row)
        print(f"  -> J1: {row['Vitorias Jogador 1']}  J2: {row['Vitorias Jogador 2']}  Empates: {row['Empates']}  Media: {row['Duracao Media do Jogo']:.3f}s")

    rows.append(placeholder("Humano vs IA (Opcional)", "Opcional no enunciado."))

    output_path = Path(args.output)
    write_results_xlsx(rows, output_path)
    print(f"\nFicheiro criado: {output_path}")


if __name__ == "__main__":
    main()
